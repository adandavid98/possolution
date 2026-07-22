import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_inventory_to_excel(products, output_filepath):
    """Exports inventory product list to an Excel (.xlsx) file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Title Block
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "MINIMARKET LA RUTA DEL ESTE - REPORTE DE INVENTARIOS"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='16161F', end_color='16161F', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["ID", "Código de Barras", "Producto", "Categoría", "Precio Costo (RD$)", "Precio Venta (RD$)", "Stock Actual", "Stock Mínimo"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color='2C2C3E', end_color='2C2C3E', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[2].height = 24

    # Data Rows
    for p in products:
        row = [
            p['id'],
            p['codigo_barras'],
            p['nombre'],
            p.get('categoria_nombre', 'N/A') or 'N/A',
            p['precio_costo'],
            p['precio_venta'],
            p['stock_actual'],
            p['stock_minimo']
        ]
        ws.append(row)

    # Styling data cells & column widths
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [5, 6]:
                cell.number_format = 'RD$#,##0.00'
            if col_idx in [1, 2, 7, 8]:
                cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_filepath)
    return output_filepath

def export_sales_to_excel(sales, output_filepath):
    """Exports sales transaction history to an Excel (.xlsx) file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Title Block
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "MINIMARKET LA RUTA DEL ESTE - REPORTE DE VENTAS"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='16161F', end_color='16161F', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ["Factura #", "Fecha", "Cliente", "Tipo Pago", "Subtotal (RD$)", "ITBIS (RD$)", "Total (RD$)"]
    ws.append(headers)

    header_fill = PatternFill(start_color='2C2C3E', end_color='2C2C3E', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 24

    for s in sales:
        ws.append([
            s['codigo_factura'],
            str(s['fecha']),
            s['cliente_nombre'],
            s['tipo_pago'],
            s['subtotal'],
            s['itbis_impuesto'],
            s['total']
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_filepath)
    return output_filepath
